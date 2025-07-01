from openpyxl import load_workbook
import requests
from selenium.webdriver import Edge
import time
from bs4 import BeautifulSoup
import os
import json
from PIL import Image
import pytesseract

file_path = "C:/website.xlsx"  # 数据文件路径
start_num = 6501  # 开始读取的序号
end_num = 6600  # 结束读取的号
output_dir = "mllm_data"  # 输出目录

# 加载工作簿（Excel文件）
wb = load_workbook(file_path)

# 选择活动的工作表，或者通过名字选择特定的工作表
sheet = wb.active  # 或者使用 wb['Sheet1'] 来指定工作表名

# 创建一个列表来存储所有的消息和图像信息
all_data = []


# 存活检测
def live_detect(url):
    if "http" in url:
        pass
    else:
        url = "http://" + url

    try:
        response = requests.get(url, timeout=5, verify=False)
        if response.status_code == 200 or response.status_code == 301 or response.status_code == 302:
            status = '存活'
        else:
            status = '失活'
    except Exception as e:
        status = '失活'
        print(e)

    return status


# 爬取信息
def pa(url, id):
    if "http" in url:
        pass
    else:
        url = "http://" + url

    driver = Edge()

    # 设置页面加载时间
    driver.set_page_load_timeout(20)
    try:
        driver.get(url)
        driver.fullscreen_window()
    except:
        return 'no'
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')

    # 提取所有文本
    all_text = soup.get_text().replace('\n', '').replace('\t','').replace(' ','')[:100]
    print(all_text)

    # 无法访问判断
    if '无法访问此页面' in all_text:
        return 'no'

    time.sleep(8)

    # 截取网页
    try:
        screenshot = driver.get_screenshot_as_png()
    except:
        return 'no'
    img_path = 'mllm_data/' + str(id) + '.png'

    # 保存截图
    with open(img_path, 'wb') as f:
        f.write(screenshot)


    #ocr提取
    try:
        # 打开图片文件
        img = Image.open(img_path)

        # 使用Tesseract进行OCR识别
        text = pytesseract.image_to_string(img, lang='chi_sim')

        return text.strip()
    except Exception as e:
        print(f"OCR处理出错: {e}")

    # 添加消息和图像信息到列表中
    all_data.append(
        {

            "messages": [
                {

                    "content": "<image>图片中为网站目标网站的网站截图\n" + all_text + text + '这是目标网站的文本内容\n请你综合这两方面信息判断该网站是否和网络赌博相关，只回答“此网站和网络赌博不相关”或“此网站和网络赌博相关”无需回答理由',
                    "role": "user",
                },
                {

                    "content": '此网站和网络赌博相关',  # 获取的文本
                    "role": "assistant",
                },
            ],
            "images": [img_path],  # 图像文件路径
        }
    )

    return 'ok'


for id in range(start_num, end_num + 1):
    print(id)
    domain = sheet.cell(row=id, column=1).value
    if "https://" in domain:  # 域名预处理
        domain = domain.split('/')[2].split(':')[0]
    elif "http://" in domain:
        domain = domain.split('/')[2].split(':')[0]
    else:
        domain = domain
    print(domain)

    if live_detect(domain) == "存活":
        status=pa(domain, id)
        if status=='no':
            continue

    # 创建 JSON 文件
    json_file_path = "mllm_data/mllm_data13.json"
    with open(json_file_path, "w", encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False)  # 确保中文字符正常显示
